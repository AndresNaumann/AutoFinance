import openpyxl as o
from openpyxl import Workbook
import os

# Ins and Outs
ins = 2
outs = 2

# Total Ins and Outs
total_ins = 0
total_outs = 0

# Essential Values
total = 0
saved = 0
play = 0

# Set Up Sheet!
if os.path.isfile('numbers.xlsx') == False:
    wb = Workbook()
    s1 = wb.active

    # Sheet Setup
    s1.title = "Data"
    s1 = wb['Data']
    wb.create_sheet(title='Vals')
    s2 = wb['Vals']

    # User inputs
    total = float(input("How much money do you already have in total? "))
    saved = float(input("How much money have you already saved? "))
    save_rate = int(input("What percentage of your money would you like to save from now on? "))
    save_rate = save_rate / 100
    play = total - saved

    tr = input("Do you pay a tithe? ")
    if tr == "yes":
        tithing = int(input("What percentage? "))
        tithing = tithing / 100

    # Organize values among sheets
    # Sheet 1 Values Setup
    s1['A1'] = "Income"
    s1['B1'] = "Expenses"
    s1['D1'] = "Total"
    s1['E1'] = total
    s1['D2'] = "Saved"
    s1['E2'] = saved
    s1['D3'] = "Leftover"
    s1["E3"] = play


    # Sheet 2 Values Setup
    s2['A1'] = "Save Rate"
    s2['B1'] = save_rate
    s2['A2'] = "Ins"
    s2['B2'] = ins
    s2['A3'] = "Outs"
    s2['B3'] = outs
    s2['A4'] = "Total In"
    s2['B4'] = total_ins
    s2['A5'] = "Total Out"
    s2['B5'] = total_outs

elif os.path.isfile('numbers.xlsx') == True:
    wb = o.load_workbook('numbers.xlsx')

    s1 = wb['Data']
    s2 = wb['Vals']

    total = s1['E1'].value
    saved = s1['E2'].value
    play = s1['E3'].value

    save_rate = s2['B1'].value
    ins = s2['B2'].value
    outs = s2['B3'].value
    total_ins = s2['B4'].value
    total_outs = s2['B5'].value

# Saves everything in the sheet
def save_sheet():
    s1['E1'] = total
    s1['E2'] = saved
    s1['E3'] = play

    s2['B2'] = ins
    s2['B3'] = outs
    s2['B4'] = total_ins
    s2['B5'] = total_outs

# User enters whether they have gained money or lost money
print("\nWelcome to the Personal Money Manager!\n")
print("Total: " + str(total) + "\nSaved: " + str(saved) + "\nPlay Money: " + str(play) + "\n")
r = "h"

while r != "no":
    wl = input("Did you gain or lose money:\n")
    amount = float(input("Enter Amount: "))

    # Things happen
    if (wl == "gain"):
        s1['A{}'.format(ins)] = amount
        total_ins = total_ins + amount
        total = total + amount
        saved = saved + total_ins * save_rate
        play = total - saved
        ins = ins + 1
        save_sheet()

    else:
        s1['B{}'.format(outs)] = amount
        total_outs = total_outs + amount
        total = total - amount
        play = total - saved
        outs = outs + 1
        save_sheet()

    r = input("Continue? ")

wb.save('numbers.xlsx')